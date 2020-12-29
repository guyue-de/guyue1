
"""
接口自动化测试
1、excel测试用例准备好，代码可以自动读取用例数据
2、执行接口测试，得到相应结果
3、响应结果和预期结果作比较，看是否通过
4、把测试是否通过的结果写入到excel工作表中
"""
#url=http://8.1729.91.152:8766/futureloan/member/register
#第一步，把读取测试用例数据封装成一个函数
import json

import requests
import openpyxl

def read_case(filename,sheetname):
    wb=openpyxl.load_workbook(filename,sheetname)#加载工作簿，打开一个Excel文件
    sheet=wb[sheetname]#打开某一个表单
    row_max=sheet.max_row#获取最大行函数
    case_list=[] #新建一个空列表，存放for循环依次读取到的测试用例
    for i in range(2,row_max+1):
        data_dict=dict(
        case_id=sheet.cell(row=i,column=1).value,
        url=sheet.cell(row=i,column=5).value,
        data=sheet.cell(row=i,column=6).value,
        expect=sheet.cell(row=i,column=7).value
        )
        case_list.append(data_dict)#把每一行读取测试用例生成的字典，逐条追加到新的列表
    return case_list
#第二部，执行接口测试
def api_fun(url,data):
    headers={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    result_register=requests.post(url=url,data=data,headers=headers).json()
    return result_register

#写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename) #加载工作簿，打开一个excel文件
    sheet=wb[sheetname]#打开某一个表单
    sheet.cell(row=row,column=column).value=final_result
    wb.save(filename)

def excute_fun(filename,sheetname):
    cases=read_case(filename, sheetname)#调用函数
    for case in cases:
        case_id=case['case_id']
        url=case['url']
        data=case['data']
        expect=eval(case['expect'])
        expect_msg=expect['msg']
        print(case_id)
        print('期望结果为{}'.format(expect_msg))
        real_result=api_fun(url=url,data=data)#调用函数
        real_msg=real_result['msg']
        print('实际结果为{}'.format(real_msg))
        if expect_msg==real_msg:
            print('这第{}条测试用例通过！'.format(case_id))
            final_result='Passed'
        else:
            print('这第{}条测试用例不通过！'.format(case_id))
            final_result='Failed'
        write_result(filename,sheetname,case_id+1,8,final_result)#调用函数
        print('*'*15)
#
# excute_fun('C:\\Users\\wzfwzf108122\\PycharmProjects\\python\\test_data\\test_case_api.xlsx','register')#调用函数
